import matplotlib.pyplot as plt, numpy as np, matplotlib
import os,sqlite3
from scipy.interpolate import interp1d
matplotlib.rc('font',**{'size':14,'family':'sans-serif'})
plt.rcParams['xtick.major.pad']='8'
plt.rcParams['ytick.major.pad']='8'
plt.rcParams['legend.fontsize']='12'
plt.rcParams['lines.markersize']='4.5'
_db_connection = sqlite3.connect('../data/iaea_xs.db')
_db = _db_connection.cursor()


class manager(object):
	def __init__(self):
		self.db_connection = _db_connection
		self.db = _db
		self.pallate = {'k':'#2c3e50','b':'#2980b9','r':'#c0392b','y':'#f39c12','p':'#8e44ad','g':'#27ae60','gy':'#7f8c8d','o':'#d35400','w':'#ecf0f1','aq':'#16a085'}
	def creat_plots_dir(self):
		os.chdir('..')
		os.system('mkdir plots')
		os.chdir('code')
	def update_exfor(self):
		files = [i[2] for i in os.walk('../data/exfor')][0]
		for istp in files:
			self.db.execute('DELETE FROM exfor WHERE isotope=?',(istp.split('.')[0],))
			for ln in open('../data/exfor/'+istp,'r').read().split('\n'):
				if not ln.startswith('#') and not ln.startswith('//'):
					d = [i.strip() for i in ln.split(' ') if i.strip()!='']
					self.db.execute('INSERT INTO exfor VALUES(?,?,?,?,?,?)',(istp.split('.')[0],float(d[0]),float(d[1]),1e3*float(d[2]),1e3*float(d[3]),d[5]))
			self.db_connection.commit()
		print 'EXFOR updated'
	def update_recommended(self):
		files = [i[2] for i in os.walk('../data/recommended')][0]
		for fnm in files:
			target,istp = fnm.split('p')[0].title(),fnm.split('p')[1].split('t.')[0].upper()
			self.db.execute('DELETE FROM monitor_xs where product=?',(istp,))
			for ln in open('../data/recommended/'+fnm,'r').read().split('\n')[5:-1]:
				if not ln.startswith(' ') and not ln.startswith('\t') and not ln.strip()=='':
					if float(ln.split('\t')[0])>100.0:
						continue
					self.db.execute('INSERT INTO monitor_xs VALUES(?,?,?,?,?)',(target,istp,float(ln.split('\t')[0]),float(ln.split('\t')[1]),0))
			self.db_connection.commit()
		print 'Recommended XS updated'
	def exp_smooth(self,ls,alpha=0.3):
		R,RR,b = [ls[0]],[ls[-1]],1.0-alpha
		for i,ii in zip(ls[1:],reversed(ls[:-1])):
			R.append(alpha*i+b*R[-1])
			RR.append(alpha*ii+b*RR[-1])
		return [0.5*(R[n]+r) for n,r in enumerate(reversed(RR))]
	def filter_outliers(self,monitor_xs,dat,mn,mx):
		dat = [d for d in dat if mn<d[0]<mx]
		return [d for d in dat if abs(d[2]-monitor_xs(d[0]))/d[3]<5 and d[3]/d[2]<0.15]
	def calculate_uncertainties(self,show=False):
		for istp in list(set([str(i[1]) for i in self.db.execute('SELECT * FROM monitor_xs')])):
			f,ax = plt.subplots()
			rec = [[float(i[2]),float(i[3]),str(i[0])] for i in self.db.execute('SELECT * FROM monitor_xs WHERE product=?',(istp,))]
			monitor = interp1d([i[0] for i in rec],[i[1] for i in rec])
			x4 = sorted(self.filter_outliers(monitor,[[float(i[1]),float(i[2]),float(i[3]),float(i[4])] for i in self.db.execute('SELECT * FROM exfor WHERE isotope=?',(istp,))],rec[0][0],rec[-1][0]),key=lambda h:h[0])
			Erange = np.arange(rec[0][0],rec[-1][0],0.1)
			E,dE,XS,dXS = [i[0] for i in x4],[i[1] for i in x4],[i[2] for i in x4],[i[3] for i in x4]
			sm_dXS = self.exp_smooth(dXS,0.1)
			sigma = interp1d(E,sm_dXS,bounds_error=False,fill_value=(sm_dXS[0],sm_dXS[-1]))
			self.db.executemany('UPDATE monitor_xs SET unc_cross_section=? WHERE product=? AND energy=?',[(float(sigma(e[0])),istp,e[0]) for e in rec])
			self.db_connection.commit()
			ax.plot(Erange,[monitor(e) for e in Erange],color=self.pallate['k'],label='IAEA Rec. XS',zorder=10)
			ax.plot(Erange,[monitor(e)+sigma(e) for e in Erange],color=self.pallate['k'],ls='--',label=r'$\pm 1\sigma$',zorder=10)
			ax.plot(Erange,[monitor(e)-sigma(e) for e in Erange],color=self.pallate['k'],ls='--',zorder=10)
			ax.errorbar(E,XS,xerr=dE,yerr=dXS,color=self.pallate['gy'],ls='None',marker='o',label='EXFOR Data',zorder=1)
			ax.set_xlabel('Energy [MeV]')
			ax.set_ylabel('Cross Section [mb]')
			ax.set_title(r'$^{nat}$'+rec[0][2]+'(p,x)'+r'$^{'+istp[:2]+r'}$'+istp[2:].title()+' Cross Section')
			ax.legend(loc=0,fontsize=10,borderaxespad=0.75)
			f.tight_layout()
			if show:
				plt.show()
			else:
				f.savefig('../plots/'+istp+'.png')
				f.savefig('../plots/'+istp+'.pdf')
				plt.close()
	def save_as_xlsx(self):
		from openpyxl import Workbook
		wb = Workbook()
		ws = wb.active
		for n,tt in enumerate(['Target','Product','E [MeV]','XS [mb]','unc_XS [mb]']):
			ws.cell(row=1,column=n+1,value=tt)
		for n,ln in enumerate([[str(i[0]),str(i[1]),float(i[2]),float(i[3]),float(i[4])] for i in self.db.execute('SELECT * FROM monitor_xs')]):
			for m,i in enumerate(ln):
				ws.cell(row=n+2,column=m+1,value=i)
		wb.save('../data/iaea_xs.xlsx')
	def save_as_csv(self):
		f = open('../data/iaea_xs.csv','w')
		ss = ','.join(['#Target','Product','E [MeV]','XS [mb]','unc_XS [mb]'])+'\n'
		f.write(ss+'\n'.join([','.join([str(i[0]),str(i[1]),str(i[2]),str(i[3]),str(i[4])]) for i in self.db.execute('SELECT * FROM monitor_xs')]))
		f.close()
	def move_data(self,target):
		db_connection = sqlite3.connect(target)
		db = db_connection.cursor()
		db.execute('DELETE FROM monitor_xs')
		db.executemany('INSERT INTO monitor_xs VALUES(?,?,?,?,?)',[(str(i[0]),str(i[1]),float(i[2]),float(i[3]),float(i[4])) for i in self.db.execute('SELECT * FROM monitor_xs')])
		db_connection.commit()


if __name__ == "__main__":
	mn = manager()
	mn.creat_plots_dir()
	mn.update_exfor()
	mn.update_recommended()
	mn.calculate_uncertainties()
	mn.save_as_csv()
	mn.save_as_xlsx()
	# mn.move_data('../../LaCe_Bernstein_Sep2017/data/peak_data.db')